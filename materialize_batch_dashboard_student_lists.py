from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


BASE_DIR = Path(__file__).resolve().parent
MASTER_FILE = BASE_DIR / "ExcelKidsHub-Master-Data.xlsx"


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def title_case(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return " ".join(part.capitalize() for part in text.split())


def main() -> None:
    workbook = load_workbook(MASTER_FILE)
    admissions_ws = workbook["admissions"]
    batches_ws = workbook["batches"]
    dashboard_ws = workbook["batch-dashboard"]

    students_by_batch: dict[str, list[str]] = defaultdict(list)
    for row in range(2, admissions_ws.max_row + 1):
        if not admissions_ws[f"A{row}"].value:
            continue
        batch_code = normalize_text(admissions_ws[f"M{row}"].value)
        student_name = title_case(admissions_ws[f"G{row}"].value)
        if batch_code and student_name:
            students_by_batch[batch_code].append(student_name)

    left_starts = [1, 7]
    card_height = 10
    row_gap = 2

    for batch_index in range(max(batches_ws.max_row - 1, 0)):
        batch_row = batch_index + 2
        batch_code = normalize_text(batches_ws[f"A{batch_row}"].value)
        if not batch_code:
            continue

        grid_row = batch_index // 2
        grid_col = batch_index % 2
        start_row = 1 + grid_row * (card_height + row_gap)
        start_col = left_starts[grid_col]
        student_list_row = start_row + 9
        student_list_cell = dashboard_ws.cell(student_list_row, start_col)

        names = students_by_batch.get(batch_code, [])
        student_list_cell.value = "\n".join(names) if names else "No students assigned"
        student_list_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    workbook.save(MASTER_FILE)
    print(f"Materialized student lists in: {MASTER_FILE}")


if __name__ == "__main__":
    main()
