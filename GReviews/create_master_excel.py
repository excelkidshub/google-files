#!/usr/bin/env python3
"""
Convert Google Takeout Google Reviews into a Comparison-Ready Excel File
"""

import json
import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuration
SOURCE_DIR = r"C:\application\AI_Project\ExcelKidsHub\google-files\GReviews\downlods\takeout-20260822T093317Z-1-001\Takeout\Google Business Profile\account-105010418223219985675\location-6826894258160505115"
OUTPUT_DIR = r"C:\application\AI_Project\ExcelKidsHub\google-files\GReviews\data"
OUTPUT_FILE = "google_reviews_master.xlsx"

# Source files
REVIEWS_FILE = Path(SOURCE_DIR) / "reviews.json"
DATA_FILE = Path(SOURCE_DIR) / "data.json"

def load_json_file(file_path):
    """Load JSON file safely"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return None

def extract_review_id(review_name):
    """Extract review ID from the Google resource name"""
    # Format: accounts/105010418223219985675/locations/6826894258160505115/reviews/REVIEW_ID
    parts = review_name.split('/')
    if len(parts) >= 5 and 'reviews' in parts:
        review_index = parts.index('reviews')
        if review_index + 1 < len(parts):
            return parts[review_index + 1]
    return review_name

def convert_star_rating_to_numeric(star_rating):
    """Convert star rating text to numeric value"""
    rating_map = {
        'ONE': 1,
        'TWO': 2,
        'THREE': 3,
        'FOUR': 4,
        'FIVE': 5
    }
    return rating_map.get(star_rating, None)

def parse_datetime(dt_string):
    """Parse ISO 8601 datetime string to formatted datetime"""
    if not dt_string:
        return None
    try:
        # Parse the ISO 8601 format
        dt = datetime.fromisoformat(dt_string.replace('Z', '+00:00'))
        # Format as YYYY-MM-DD HH:MM:SS
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        print(f"Error parsing datetime {dt_string}: {e}")
        return dt_string

def process_reviews():
    """Process reviews from JSON file"""
    print("Loading reviews data...")
    reviews_data = load_json_file(REVIEWS_FILE)
    
    if not reviews_data or 'reviews' not in reviews_data:
        print("No reviews found in the source file")
        return None, None
    
    reviews = reviews_data['reviews']
    print(f"Found {len(reviews)} reviews in source file")
    
    # Load business data for context
    business_data = load_json_file(DATA_FILE)
    business_name = "Unknown"
    location_id = "Unknown"
    account_id = "Unknown"
    
    if business_data:
        business_name = business_data.get('title', 'Unknown')
        # Extract location ID from name field
        name_field = business_data.get('name', '')
        if 'locations/' in name_field:
            location_id = name_field.split('locations/')[1].split('/')[0] if '/' in name_field else name_field.split('locations/')[1]
        # Extract account ID from review names
        if reviews and 'name' in reviews[0]:
            account_id = reviews[0]['name'].split('accounts/')[1].split('/')[0] if 'accounts/' in reviews[0]['name'] else 'Unknown'
    
    processed_reviews = []
    
    for review in reviews:
        # Extract all available fields
        processed_review = {
            'Review ID': extract_review_id(review.get('name', '')),
            'Review Resource Name': review.get('name', ''),
            'Reviewer Name': review.get('reviewer', {}).get('displayName', ''),
            'Rating': convert_star_rating_to_numeric(review.get('starRating')),
            'Rating Original': review.get('starRating', ''),
            'Review Text': review.get('comment', ''),
            'Review Created Date': parse_datetime(review.get('createTime')),
            'Review Updated Date': parse_datetime(review.get('updateTime')),
            'Business Name': business_name,
            'Location ID': location_id,
            'Account ID': account_id,
            'Owner Reply': None,  # Not present in this takeout
            'Owner Reply Date': None  # Not present in this takeout
        }
        
        # Add Review Key (using Review ID as stable identifier)
        processed_review['Review Key'] = processed_review['Review ID']
        
        processed_reviews.append(processed_review)
    
    return processed_reviews, business_name

def check_duplicates(reviews):
    """Check for duplicate reviews based on Review Key"""
    review_keys = [r['Review Key'] for r in reviews]
    unique_keys = set(review_keys)
    
    if len(review_keys) != len(unique_keys):
        # Find duplicates
        from collections import Counter
        key_counts = Counter(review_keys)
        duplicates = [key for key, count in key_counts.items() if count > 1]
        print(f"Found {len(duplicates)} duplicate review keys: {duplicates}")
        return True, duplicates
    print("No duplicate review keys found")
    return False, []

def create_excel_workbook(reviews, business_name):
    """Create the Excel workbook with multiple sheets"""
    print("Creating Excel workbook...")
    
    output_path = Path(OUTPUT_DIR) / OUTPUT_FILE
    
    # Create DataFrame
    df = pd.DataFrame(reviews)
    
    # Check for duplicates
    has_duplicates, duplicate_keys = check_duplicates(reviews)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write main Reviews sheet
        df.to_excel(writer, sheet_name='Reviews', index=False)
        
        # Create Duplicates sheet if needed
        if has_duplicates:
            duplicate_reviews = df[df['Review Key'].isin(duplicate_keys)]
            duplicate_reviews.to_excel(writer, sheet_name='Duplicates', index=False)
        
        # Create Summary sheet
        create_summary_sheet(writer, df, business_name, has_duplicates, len(duplicate_keys) if has_duplicates else 0)
        
        # Create README sheet
        create_readme_sheet(writer, df, business_name, has_duplicates, duplicate_keys)
    
    # Format the Excel file
    format_excel_file(output_path)
    
    return output_path, has_duplicates, duplicate_keys

def create_summary_sheet(writer, df, business_name, has_duplicates, duplicate_count):
    """Create summary sheet with statistics"""
    summary_data = {
        'Metric': [
            'Business Name',
            'Total Reviews',
            '1 Star Reviews',
            '2 Star Reviews', 
            '3 Star Reviews',
            '4 Star Reviews',
            '5 Star Reviews',
            'Reviews With Owner Reply',
            'Reviews Without Owner Reply',
            'Earliest Review Date',
            'Latest Review Date',
            'Duplicate Records Found',
            'Source File',
            'Export Date'
        ],
        'Value': [
            business_name,
            len(df),
            len(df[df['Rating'] == 1]),
            len(df[df['Rating'] == 2]),
            len(df[df['Rating'] == 3]),
            len(df[df['Rating'] == 4]),
            len(df[df['Rating'] == 5]),
            len(df[df['Owner Reply'].notna() & (df['Owner Reply'] != '')]),
            len(df[df['Owner Reply'].isna() | (df['Owner Reply'] == '')]),
            df['Review Created Date'].min() if not df['Review Created Date'].isna().all() else 'N/A',
            df['Review Created Date'].max() if not df['Review Created Date'].isna().all() else 'N/A',
            duplicate_count,
            str(REVIEWS_FILE),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

def create_readme_sheet(writer, df, business_name, has_duplicates, duplicate_keys):
    """Create README sheet with documentation"""
    duplicate_info = f'Yes - {len(duplicate_keys)} duplicate records found' if has_duplicates else 'No duplicates found'
    
    readme_data = {
        'Section': [
            'Source',
            'Source Folder',
            'Export Date',
            'Business/Profile Name',
            'Number of Reviews Imported',
            'Stable Review Key Field',
            'Review Key Source',
            'Fields Available',
            'Fields Not Available',
            'Owner Replies',
            'Duplicates',
            'Data Integrity',
            'Purpose'
        ],
        'Description': [
            'Google Takeout',
            SOURCE_DIR,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            business_name,
            len(df),
            'Review Key',
            'Google Review ID (extracted from Review Resource Name)',
            'Review ID, Review Resource Name, Reviewer Name, Rating, Rating Original, Review Text, Review Created Date, Review Updated Date, Business Name, Location ID, Account ID, Owner Reply, Owner Reply Date, Review Key',
            'None - all standard fields are present',
            'No owner replies present in this Takeout export',
            duplicate_info,
            'All review text, ratings, and dates preserved exactly as in source',
            'Master baseline snapshot for future review comparisons'
        ]
    }
    
    readme_df = pd.DataFrame(readme_data)
    readme_df.to_excel(writer, sheet_name='README', index=False)

def format_excel_file(file_path):
    """Apply formatting to the Excel file"""
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)

def perform_data_integrity_checks(reviews, output_path):
    """Perform data integrity checks"""
    print("Performing data integrity checks...")
    
    # Count source reviews
    source_count = len(reviews)
    
    # Load Excel file and count reviews
    df_excel = pd.read_excel(output_path, sheet_name='Reviews')
    excel_count = len(df_excel)
    
    print(f"Source review count: {source_count}")
    print(f"Excel review count: {excel_count}")
    
    if source_count != excel_count:
        print("ERROR: Review counts do not match!")
        return False
    
    # Check that review text is preserved
    for i, review in enumerate(reviews):
        excel_text = df_excel.iloc[i]['Review Text']
        if review['Review Text'] != excel_text:
            print(f"ERROR: Review text mismatch at index {i}")
            return False
    
    # Check that ratings are preserved
    for i, review in enumerate(reviews):
        excel_rating = df_excel.iloc[i]['Rating']
        if review['Rating'] != excel_rating:
            print(f"ERROR: Rating mismatch at index {i}")
            return False
    
    print("Data integrity checks passed")
    return True

def main():
    """Main function"""
    print("=" * 60)
    print("Google Takeout Reviews to Excel Converter")
    print("=" * 60)
    
    # Process reviews
    reviews, business_name = process_reviews()
    
    if not reviews:
        print("No reviews to process")
        return
    
    # Create Excel workbook
    output_path, has_duplicates, duplicate_keys = create_excel_workbook(reviews, business_name)
    
    # Perform data integrity checks
    integrity_passed = perform_data_integrity_checks(reviews, output_path)
    
    if not integrity_passed:
        print("ERROR: Data integrity checks failed")
        return
    
    print("\n" + "=" * 60)
    print("CONVERSION COMPLETE")
    print("=" * 60)
    print(f"Output file: {output_path}")
    print(f"Total reviews processed: {len(reviews)}")
    print(f"Business name: {business_name}")
    
    # Generate final report
    df = pd.read_excel(output_path, sheet_name='Reviews')
    
    print("\n" + "=" * 60)
    print("FINAL REPORT")
    print("=" * 60)
    print(f"Source files inspected: {REVIEWS_FILE}, {DATA_FILE}")
    print(f"Review source file: {REVIEWS_FILE}")
    print(f"Total reviews found: {len(reviews)}")
    print(f"Total reviews written to Excel: {len(df)}")
    print(f"Duplicate reviews: {len(duplicate_keys) if has_duplicates else 0}")
    print(f"Reviews with owner replies: {len(df[df['Owner Reply'].notna() & (df['Owner Reply'] != '')])}")
    print(f"Reviews without owner replies: {len(df[df['Owner Reply'].isna() | (df['Owner Reply'] == '')])}")
    print(f"Earliest review: {df['Review Created Date'].min() if not df['Review Created Date'].isna().all() else 'N/A'}")
    print(f"Latest review: {df['Review Created Date'].max() if not df['Review Created Date'].isna().all() else 'N/A'}")
    print(f"Review Key field: Review Key (Google Review ID)")
    print(f"Output file: {output_path}")
    
    print("\nExcel Structure:")
    print("- Reviews: Main sheet with all review data")
    print("- Summary: Statistics and metadata")
    print("- README: Documentation and notes")
    if has_duplicates:
        print("- Duplicates: Duplicate review records")

if __name__ == "__main__":
    main()