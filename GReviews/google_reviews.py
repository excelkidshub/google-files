#!/usr/bin/env python3
"""
Google Business Profile Reviews Backup System
Downloads and tracks Google Business Profile reviews with historical comparison.
"""

import os
import json
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import sys
import ssl

import pandas as pd
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Fix SSL certificate issue on Windows
ssl._create_default_https_context = ssl._create_unverified_context


class GoogleReviewsBackup:
    """Main class for Google Business Profile reviews backup and tracking."""
    
    def __init__(self, config_path: str = 'config.json'):
        self.config = self._load_config(config_path)
        self._setup_directories()
        self._setup_logging()
        self.current_date = datetime.now()
        
    def _load_config(self, config_path: str) -> dict:
        """Load configuration from JSON file."""
        with open(config_path, 'r') as f:
            return json.load(f)
    
    def _setup_directories(self):
        """Ensure all required directories exist."""
        for dir_path in [
            self.config['paths']['credentials_dir'],
            self.config['paths']['data_dir'],
            self.config['paths']['backup_dir'],
            self.config['paths']['snapshots_dir'],
            self.config['paths']['reports_dir'],
            self.config['paths']['logs_dir']
        ]:
            Path(dir_path).mkdir(parents=True, exist_ok=True)
    
    def _setup_logging(self):
        """Configure logging system."""
        log_file = self.config['paths']['log_file']
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def _authenticate(self) -> Credentials:
        """Authenticate with Google OAuth 2.0."""
        creds = None
        token_path = self.config['api']['token_file']
        credentials_path = self.config['api']['credentials_file']
        scopes = self.config['api']['scopes']
        
        if os.path.exists(token_path):
            with open(token_path, 'r') as f:
                creds_data = json.load(f)
            creds = Credentials.from_authorized_user_info(creds_data, scopes)
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                self.logger.info("Refreshing expired token...")
                creds.refresh(Request())
            else:
                if not os.path.exists(credentials_path):
                    raise FileNotFoundError(
                        f"Credentials file not found: {credentials_path}\n"
                        "Please follow the setup instructions in credentials/README.txt"
                    )
                self.logger.info("Starting OAuth authentication flow...")
                flow = InstalledAppFlow.from_client_secrets_file(credentials_path, scopes)
                creds = flow.run_local_server(port=0)
            
            with open(token_path, 'w') as f:
                json.dump(json.loads(creds.to_json()), f, indent=2)
        
        return creds
    
    def _diagnose_api_access(self, creds: Credentials):
        """Diagnose API access for account, location, and review endpoints."""
        print("=" * 60)
        print("API ACCESS DIAGNOSTIC")
        print("=" * 60)
        print()
        
        # Test 1: Account Listing (My Business Account Management API)
        print("Test 1: Account Listing (My Business Account Management API)")
        print("-" * 60)
        try:
            mybusiness_service = build('mybusinessaccountmanagement', 'v1', credentials=creds)
            accounts_response = mybusiness_service.accounts().list().execute()
            accounts = accounts_response.get('accounts', [])
            print(f"✓ SUCCESS: Found {len(accounts)} account(s)")
            for account in accounts:
                print(f"  - {account.get('name', 'Unknown')}")
        except HttpError as e:
            if e.resp.status == 429:
                print(f"✗ FAILED: Rate limit exceeded (Quota: 0 requests/minute)")
                print(f"  Error: {e}")
            elif e.resp.status == 403:
                print(f"✗ FAILED: Permission denied")
                print(f"  Error: {e}")
            else:
                print(f"✗ FAILED: HTTP {e.resp.status}")
                print(f"  Error: {e}")
        except Exception as e:
            print(f"✗ FAILED: {e}")
        print()
        
        # Test 2: Location Listing (My Business Account Management API)
        account_id = self.config['google_business_profile']['account_id']
        print(f"Test 2: Location Listing for account {account_id}")
        print("-" * 60)
        if not account_id:
            print("⊘ SKIPPED: No account_id in config.json")
        else:
            try:
                mybusiness_service = build('mybusinessaccountmanagement', 'v1', credentials=creds)
                account_name = f"accounts/{account_id}"
                locations_response = mybusiness_service.accounts().listLocations(parent=account_name, pageSize=100).execute()
                locations = locations_response.get('locations', [])
                print(f"✓ SUCCESS: Found {len(locations)} location(s)")
                for location in locations:
                    loc_name = location.get('locationName', 'Unknown')
                    loc_id = location.get('name', '').split('/')[-1] if '/' in location.get('name', '') else location.get('name', '')
                    print(f"  - {loc_name} (ID: {loc_id})")
            except HttpError as e:
                if e.resp.status == 429:
                    print(f"✗ FAILED: Rate limit exceeded (Quota: 0 requests/minute)")
                    print(f"  Error: {e}")
                elif e.resp.status == 403:
                    print(f"✗ FAILED: Permission denied")
                    print(f"  Error: {e}")
                else:
                    print(f"✗ FAILED: HTTP {e.resp.status}")
                    print(f"  Error: {e}")
            except Exception as e:
                print(f"✗ FAILED: {e}")
        print()
        
        # Test 3: Review Listing (Business Profile Performance API)
        location_id = self.config['google_business_profile']['location_id']
        print(f"Test 3: Review Listing for location {location_id}")
        print("-" * 60)
        if not location_id:
            print("⊘ SKIPPED: No location_id in config.json")
        else:
            try:
                performance_service = build('businessprofileperformance', 'v1', credentials=creds)
                reviews_response = performance_service.locations().fetchReviews(
                    parent=f"locations/{location_id}",
                    pageSize=1
                ).execute()
                reviews = reviews_response.get('reviews', [])
                print(f"✓ SUCCESS: Found {len(reviews)} review(s) (showing first 1)")
            except HttpError as e:
                if e.resp.status == 429:
                    print(f"✗ FAILED: Rate limit exceeded")
                    print(f"  Error: {e}")
                elif e.resp.status == 403:
                    print(f"✗ FAILED: Permission denied")
                    print(f"  Error: {e}")
                elif e.resp.status == 404:
                    print(f"✗ FAILED: Location not found")
                    print(f"  Error: {e}")
                else:
                    print(f"✗ FAILED: HTTP {e.resp.status}")
                    print(f"  Error: {e}")
            except Exception as e:
                print(f"✗ FAILED: {e}")
        print()
        
        print("=" * 60)
        print("DIAGNOSTIC COMPLETE")
        print("=" * 60)
    
    def _discover_location_for_account(self, creds: Credentials, account_id: str) -> str:
        """Discover location_id for a known account_id."""
        print(f"Discovering locations for account: {account_id}")
        print()
        
        mybusiness_service = build('mybusinessaccountmanagement', 'v1', credentials=creds)
        account_name = f"accounts/{account_id}"
        target_location_name = self.config['google_business_profile']['location_name']
        
        try:
            # Try using the list method to get accounts with locations
            accounts_response = mybusiness_service.accounts().list().execute()
            accounts = accounts_response.get('accounts', [])
            
            print(f"Found {len(accounts)} account(s) in total")
            
            # Find our specific account
            target_account = None
            for account in accounts:
                print(f"Account: {account.get('name')}")
                if account.get('name') == account_name:
                    target_account = account
                    print(f"  -> This is our target account")
                    print(f"  Account data keys: {list(account.keys())}")
            
            if not target_account:
                raise Exception(f"Account {account_id} not found in list of accounts")
            
            # Check if locations are included in the account response
            locations = target_account.get('locations', [])
            
            if not locations:
                raise Exception(f"No locations found for account {account_id}. The API may not return locations in the account list. You may need to manually enter the location_id in config.json.")
            
            print(f"Found {len(locations)} location(s) in this account:")
            for i, location in enumerate(locations, 1):
                loc_name = location.get('locationName', 'Unknown')
                loc_id = location.get('name', '').split('/')[-1] if '/' in location.get('name', '') else location.get('name', '')
                print(f"  {i}. {loc_name} (ID: {loc_id})")
            print()
            
            # Auto-select based on location_name
            matching_locations = [loc for loc in locations if target_location_name.lower() in loc.get('locationName', '').lower()]
            
            if matching_locations:
                selected_location = matching_locations[0]
                print(f"Automatically selected: {selected_location.get('locationName', 'Unknown')} (matches '{target_location_name}')")
            else:
                # No match, use first location
                selected_location = locations[0]
                print(f"No match for '{target_location_name}'. Automatically selected: {selected_location.get('locationName', 'Unknown')}")
            
            location_name = selected_location.get('locationName', 'Unknown')
            location_id = selected_location.get('name', '').split('/')[-1] if '/' in selected_location.get('name', '') else selected_location.get('name', '')
            print(f"Selected location: {location_name}")
            print()
            
            # Save to config
            self._save_config(account_id, location_id, location_name)
            
            return location_id
            
        except HttpError as e:
            if e.resp.status == 429:
                raise Exception(
                    "Rate limit exceeded for My Business Account Management API.\n"
                    "Your Google Cloud project has a quota of 0 requests per minute.\n"
                    "To fix this:\n"
                    "1. Go to https://console.cloud.google.com/apis/api/mybusinessaccountmanagement.googleapis.com/quotas\n"
                    "2. Request a quota increase for 'Requests per minute'\n"
                    "3. Alternatively, manually enter your location_id in config.json"
                )
            elif e.resp.status == 403:
                raise Exception(
                    "Permission denied. Please ensure:\n"
                    "1. You have enabled the 'My Business Account Management API' in Google Cloud Console\n"
                    "2. Your OAuth client has the correct permissions\n"
                    "3. Your Google account has access to Google Business Profile"
                )
            else:
                raise Exception(f"Google API error while discovering locations: {e}")
        except Exception as e:
            raise Exception(f"Error discovering location: {e}")
    
    def _save_config(self, account_id: str, location_id: str, location_name: str):
        """Save discovered account and location IDs to config.json."""
        self.config['google_business_profile']['account_id'] = account_id
        self.config['google_business_profile']['location_id'] = location_id
        self.config['google_business_profile']['location_name'] = location_name
        
        with open('config.json', 'w') as f:
            json.dump(self.config, f, indent=2)
        
        self.logger.info(f"Saved configuration: account_id={account_id}, location_id={location_id}")
        print(f"Configuration saved to config.json")
        print()
    
    def _get_google_reviews(self) -> List[dict]:
        """Download all reviews from Google Business Profile API."""
        creds = self._authenticate()
        
        # Check if we need to discover location_id
        account_id = self.config['google_business_profile']['account_id']
        location_id = self.config['google_business_profile']['location_id']
        location_name = self.config['google_business_profile']['location_name']
        skip_discovery = self.config['google_business_profile'].get('skip_location_discovery', False)
        
        if not location_id and not skip_discovery:
            if account_id:
                # We have account_id, discover location_id
                print("Location ID not found in configuration.")
                print("Starting location discovery using known account_id...")
                print()
                try:
                    location_id = self._discover_location_for_account(creds, account_id)
                    location_name = self.config['google_business_profile']['location_name']
                except Exception as e:
                    print(f"Location discovery failed: {e}")
                    print()
                    print("To fix this issue:")
                    print("1. Go to https://console.cloud.google.com/apis/api/mybusinessaccountmanagement.googleapis.com/quotas")
                    print("2. Request a quota increase for 'Requests per minute'")
                    print("3. OR: Find your location_id manually and add it to config.json")
                    print("4. OR: Set 'skip_location_discovery': true in config.json and provide location_id")
                    raise Exception("Location discovery failed due to API quota limits. Please add location_id to config.json manually.")
            else:
                # No account_id either, cannot proceed
                raise Exception(
                    "Both account_id and location_id are missing from config.json.\n"
                    "Please enter your account_id in config.json to enable automatic location discovery,\n"
                    "or enter both account_id and location_id manually."
                )
        elif not location_id and skip_discovery:
            raise Exception(
                "skip_location_discovery is set to true but location_id is missing from config.json.\n"
                "Please add your location_id to config.json."
            )
        
        # Verify IDs before retrieving reviews
        print(f"Account ID: {account_id}")
        print(f"Location ID: {location_id}")
        print(f"Location Name: {location_name}")
        print()
        
        # Use the My Business Account Management API (already enabled) to get locations with reviews
        service = build('mybusinessaccountmanagement', 'v1', credentials=creds)
        
        self.logger.info(f"Connecting to Google Business Profile: {location_name}")
        self.logger.info(f"Account ID: {account_id}, Location ID: {location_id}")
        
        reviews = []
        page_token = None
        page_count = 0
        
        try:
            while True:
                page_count += 1
                self.logger.info(f"Downloading reviews - Page {page_count}...")
                
                # Use the list method with readMask to include reviews
                request = service.accounts().locations().list(
                    parent=f"accounts/{account_id}",
                    pageSize=100,
                    readMask="name,locationName,reviews",
                    pageToken=page_token
                )
                
                response = request.execute()
                
                if 'locations' in response:
                    locations_data = response['locations']
                    for location in locations_data:
                        # Extract reviews from location if present
                        if 'reviews' in location:
                            reviews.extend(location['reviews'])
                    
                    self.logger.info(f"  Retrieved {len(reviews)} total reviews so far")
                
                page_token = response.get('nextPageToken')
                if not page_token:
                    break
                    
        except HttpError as e:
            self.logger.error(f"Google API error: {e}")
            raise Exception(f"Google review retrieval failed: {e}")
        except Exception as e:
            self.logger.error(f"Error downloading reviews: {e}")
            raise
        
        self.logger.info(f"Successfully downloaded {len(reviews)} total reviews")
        return reviews
    
    def _normalize_review(self, review: dict) -> dict:
        """Normalize a Google API review to a standard format."""
        review_id = review.get('name', '').split('/')[-1]
        
        # Extract review data
        normalized = {
            'review_id': review_id,
            'reviewer_name': review.get('reviewer', {}).get('displayName', ''),
            'rating': review.get('starRating', {}).get('rating', ''),
            'review_text': review.get('comment', ''),
            'creation_date': self._parse_google_date(review.get('createTime')),
            'update_date': self._parse_google_date(review.get('updateTime')),
            'owner_reply': review.get('reviewReply', {}).get('comment', ''),
            'owner_reply_date': self._parse_google_date(review.get('reviewReply', {}).get('updateTime')),
            'review_url': f"https://search.google.com/local/writereview?placeid={review_id}"
        }
        
        return normalized
    
    def _parse_google_date(self, date_str: Optional[str]) -> str:
        """Parse Google API date format to readable format."""
        if not date_str:
            return ''
        try:
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return date_str
    
    def _load_historical_data(self) -> pd.DataFrame:
        """Load historical review data from Excel file."""
        excel_path = self.config['paths']['main_excel_file']
        
        if not os.path.exists(excel_path):
            return pd.DataFrame()
        
        try:
            # Try to read the Review History sheet
            df = pd.read_excel(excel_path, sheet_name='Review History')
            self.logger.info(f"Loaded {len(df)} historical records")
            return df
        except Exception as e:
            self.logger.warning(f"Could not load historical data: {e}")
            return pd.DataFrame()
    
    def _backup_historical_file(self):
        """Create a timestamped backup of the historical Excel file."""
        excel_path = self.config['paths']['main_excel_file']
        
        if not os.path.exists(excel_path):
            return
        
        backup_dir = self.config['paths']['backup_dir']
        timestamp = self.current_date.strftime('%Y-%m-%d_%H%M%S')
        backup_filename = f"google_reviews_backup_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        shutil.copy2(excel_path, backup_path)
        self.logger.info(f"Created backup: {backup_path}")
    
    def _create_snapshot(self):
        """Create a timestamped snapshot of the current Excel file."""
        excel_path = self.config['paths']['main_excel_file']
        
        if not os.path.exists(excel_path):
            return
        
        snapshots_dir = self.config['paths']['snapshots_dir']
        timestamp = self.current_date.strftime('%Y-%m-%d')
        snapshot_filename = f"google_reviews_{timestamp}.xlsx"
        snapshot_path = os.path.join(snapshots_dir, snapshot_filename)
        
        # Only create one snapshot per day
        if not os.path.exists(snapshot_path):
            shutil.copy2(excel_path, snapshot_path)
            self.logger.info(f"Created snapshot: {snapshot_path}")
    
    def _compare_reviews(self, current_reviews: List[dict], historical_df: pd.DataFrame) -> dict:
        """Compare current reviews with historical data."""
        current_ids = set(r['review_id'] for r in current_reviews)
        
        if historical_df.empty:
            # First run - no comparison needed
            return {
                'is_first_run': True,
                'new_reviews': current_reviews,
                'existing_reviews': [],
                'missing_reviews': [],
                'returned_reviews': [],
                'review_changes': []
            }
        
        historical_ids = set(historical_df['review_id'].tolist())
        
        # Find new reviews
        new_reviews = [r for r in current_reviews if r['review_id'] not in historical_ids]
        
        # Find existing reviews
        existing_reviews = [r for r in current_reviews if r['review_id'] in historical_ids]
        
        # Find missing reviews (in history but not in current)
        missing_ids = historical_ids - current_ids
        missing_reviews = historical_df[historical_df['review_id'].isin(missing_ids)].to_dict('records')
        
        # Find returned reviews (previously missing, now present)
        previously_missing = historical_df[historical_df['current_status'] == 'MISSING']
        returned_ids = set(previously_missing['review_id'].tolist()) & current_ids
        returned_reviews = [r for r in current_reviews if r['review_id'] in returned_ids]
        
        # Detect review changes (edits)
        review_changes = []
        for current_review in existing_reviews:
            hist_record = historical_df[historical_df['review_id'] == current_review['review_id']].iloc[0]
            
            changes = []
            if current_review['review_text'] != hist_record.get('review_text', ''):
                changes.append({
                    'field': 'review_text',
                    'old_value': hist_record.get('review_text', ''),
                    'new_value': current_review['review_text']
                })
            if current_review['rating'] != hist_record.get('rating', ''):
                changes.append({
                    'field': 'rating',
                    'old_value': hist_record.get('rating', ''),
                    'new_value': current_review['rating']
                })
            
            if changes:
                review_changes.append({
                    'review_id': current_review['review_id'],
                    'reviewer_name': current_review['reviewer_name'],
                    'changes': changes,
                    'change_date': self.current_date.strftime('%Y-%m-%d %H:%M:%S')
                })
        
        return {
            'is_first_run': False,
            'new_reviews': new_reviews,
            'existing_reviews': existing_reviews,
            'missing_reviews': missing_reviews,
            'returned_reviews': returned_reviews,
            'review_changes': review_changes
        }
    
    def _update_historical_data(self, current_reviews: List[dict], comparison: dict) -> pd.DataFrame:
        """Update historical data with current reviews."""
        historical_df = self._load_historical_data()
        current_date_str = self.current_date.strftime('%Y-%m-%d')
        
        if historical_df.empty:
            # First run - create initial historical records
            historical_records = []
            for review in current_reviews:
                historical_records.append({
                    'review_id': review['review_id'],
                    'reviewer_name': review['reviewer_name'],
                    'rating': review['rating'],
                    'review_text': review['review_text'],
                    'creation_date': review['creation_date'],
                    'update_date': review['update_date'],
                    'owner_reply': review['owner_reply'],
                    'owner_reply_date': review['owner_reply_date'],
                    'first_seen': current_date_str,
                    'last_seen': current_date_str,
                    'current_status': 'PRESENT',
                    'times_seen': 1
                })
            return pd.DataFrame(historical_records)
        
        # Update existing historical data
        current_ids = set(r['review_id'] for r in current_reviews)
        
        # Update records for present reviews
        for review in current_reviews:
            idx = historical_df.index[historical_df['review_id'] == review['review_id']]
            
            if len(idx) > 0:
                # Update existing record
                idx = idx[0]
                historical_df.at[idx, 'last_seen'] = current_date_str
                historical_df.at[idx, 'current_status'] = 'PRESENT'
                historical_df.at[idx, 'times_seen'] += 1
                historical_df.at[idx, 'review_text'] = review['review_text']
                historical_df.at[idx, 'rating'] = review['rating']
                historical_df.at[idx, 'update_date'] = review['update_date']
                historical_df.at[idx, 'owner_reply'] = review['owner_reply']
                historical_df.at[idx, 'owner_reply_date'] = review['owner_reply_date']
            else:
                # Add new record
                new_record = {
                    'review_id': review['review_id'],
                    'reviewer_name': review['reviewer_name'],
                    'rating': review['rating'],
                    'review_text': review['review_text'],
                    'creation_date': review['creation_date'],
                    'update_date': review['update_date'],
                    'owner_reply': review['owner_reply'],
                    'owner_reply_date': review['owner_reply_date'],
                    'first_seen': current_date_str,
                    'last_seen': current_date_str,
                    'current_status': 'PRESENT',
                    'times_seen': 1
                }
                historical_df = pd.concat([historical_df, pd.DataFrame([new_record])], ignore_index=True)
        
        # Mark missing reviews
        missing_ids = set(historical_df['review_id'].tolist()) - current_ids
        for review_id in missing_ids:
            idx = historical_df.index[historical_df['review_id'] == review_id][0]
            if historical_df.at[idx, 'current_status'] != 'MISSING':
                historical_df.at[idx, 'current_status'] = 'MISSING'
        
        return historical_df
    
    def _create_excel_report(self, current_reviews: List[dict], historical_df: pd.DataFrame, comparison: dict):
        """Create Excel report with all required sheets."""
        excel_path = self.config['paths']['main_excel_file']
        
        # Create Excel writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Current Reviews
            current_df = pd.DataFrame(current_reviews)
            current_df.to_excel(writer, sheet_name='Current Reviews', index=False)
            
            # Sheet 2: Review History
            historical_df.to_excel(writer, sheet_name='Review History', index=False)
            
            # Sheet 3: Missing Reviews
            if comparison['missing_reviews']:
                missing_df = pd.DataFrame(comparison['missing_reviews'])
                # Add calculated columns
                current_date = self.current_date
                missing_df['days_since_last_seen'] = missing_df['last_seen'].apply(
                    lambda x: (current_date - datetime.strptime(x, '%Y-%m-%d')).days if x else ''
                )
                missing_df.to_excel(writer, sheet_name='Missing Reviews', index=False)
            else:
                # Create empty sheet with headers
                empty_missing = pd.DataFrame(columns=[
                    'review_id', 'reviewer_name', 'rating', 'creation_date',
                    'review_text', 'first_seen', 'last_seen', 'current_status',
                    'times_seen', 'days_since_last_seen'
                ])
                empty_missing.to_excel(writer, sheet_name='Missing Reviews', index=False)
            
            # Sheet 4: Run Summary
            self._create_summary_sheet(writer, current_reviews, historical_df, comparison)
            
            # Sheet 5: Review Changes (optional)
            if comparison['review_changes']:
                changes_data = []
                for change in comparison['review_changes']:
                    for ch in change['changes']:
                        changes_data.append({
                            'review_id': change['review_id'],
                            'reviewer_name': change['reviewer_name'],
                            'field': ch['field'],
                            'old_value': ch['old_value'],
                            'new_value': ch['new_value'],
                            'change_date': change['change_date']
                        })
                changes_df = pd.DataFrame(changes_data)
                changes_df.to_excel(writer, sheet_name='Review Changes', index=False)
        
        # Apply formatting
        self._apply_excel_formatting(excel_path)
        self.logger.info(f"Excel report created: {excel_path}")
    
    def _create_summary_sheet(self, writer, current_reviews: List[dict], historical_df: pd.DataFrame, comparison: dict):
        """Create the Run Summary sheet."""
        current_date_str = self.current_date.strftime('%d-%b-%Y')
        
        if comparison['is_first_run']:
            status = "BASELINE CREATED"
        else:
            status = "COMPARISON COMPLETED"
        
        summary_data = {
            'Metric': [
                'Run Date',
                'Location',
                'Current reviews found',
                'New reviews',
                'Previously known reviews',
                'Potentially missing reviews',
                'Returned reviews',
                'Total historical records',
                'Run status'
            ],
            'Value': [
                current_date_str,
                self.config['google_business_profile']['location_name'],
                len(current_reviews),
                len(comparison['new_reviews']),
                len(comparison['existing_reviews']),
                len(comparison['missing_reviews']),
                len(comparison['returned_reviews']),
                len(historical_df),
                status
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Run Summary', index=False)
    
    def _apply_excel_formatting(self, excel_path: str):
        """Apply formatting to Excel sheets."""
        wb = openpyxl.load_workbook(excel_path)
        
        # Format Missing Reviews sheet
        if 'Missing Reviews' in wb.sheetnames:
            ws = wb['Missing Reviews']
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Auto-filter
            ws.auto_filter.ref = ws.dimensions
            
            # Set column widths
            column_widths = {
                'A': 20,  # review_id
                'B': 25,  # reviewer_name
                'C': 10,  # rating
                'D': 20,  # creation_date
                'E': 50,  # review_text
                'F': 15,  # first_seen
                'G': 15,  # last_seen
                'H': 15,  # current_status
                'I': 15,  # times_seen
                'J': 20   # days_since_last_seen
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Format header row
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight MISSING status
            for row in ws.iter_rows(min_row=2):
                status_cell = row[7]  # Column H (current_status)
                if status_cell.value == 'MISSING':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(bold=True, color="C00000")
                
                # Wrap text for review content
                review_text_cell = row[4]  # Column E (review_text)
                review_text_cell.alignment = Alignment(wrap_text=True)
        
        # Format Current Reviews sheet
        if 'Current Reviews' in wb.sheetnames:
            ws = wb['Current Reviews']
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
            
            column_widths = {
                'A': 20, 'B': 25, 'C': 10, 'D': 50, 'E': 20,
                'F': 20, 'G': 50, 'H': 20, 'I': 30, 'J': 15, 'K': 15, 'L': 15
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        # Format Review History sheet
        if 'Review History' in wb.sheetnames:
            ws = wb['Review History']
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
            
            column_widths = {
                'A': 20, 'B': 25, 'C': 10, 'D': 50, 'E': 20,
                'F': 20, 'G': 50, 'H': 20, 'I': 15, 'J': 15, 'K': 15, 'L': 15
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        # Format Run Summary sheet
        if 'Run Summary' in wb.sheetnames:
            ws = wb['Run Summary']
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        wb.save(excel_path)
    
    def run(self):
        """Main execution method."""
        print("=" * 40)
        print("ExcelKidsHub Google Reviews Backup")
        print("=" * 40)
        print()
        
        try:
            # Step 1: Download current reviews from Google
            print("Connecting to Google Business Profile...")
            current_reviews_raw = self._get_google_reviews()
            
            # Step 2: Normalize reviews
            print("Normalizing review data...")
            current_reviews = [self._normalize_review(r) for r in current_reviews_raw]
            
            # Step 3: Load historical data
            print("Loading historical records...")
            historical_df = self._load_historical_data()
            
            # Step 4: Backup existing file if it exists
            if not historical_df.empty:
                self._backup_historical_file()
            
            # Step 5: Compare reviews
            print("Comparing reviews...")
            comparison = self._compare_reviews(current_reviews, historical_df)
            
            # Step 6: Update historical data
            print("Updating historical database...")
            updated_historical_df = self._update_historical_data(current_reviews, comparison)
            
            # Step 7: Create Excel report
            print("Creating Excel report...")
            self._create_excel_report(current_reviews, updated_historical_df, comparison)
            
            # Step 8: Create snapshot
            self._create_snapshot()
            
            # Display results
            print()
            print("=" * 40)
            
            if comparison['is_first_run']:
                print("FIRST RUN DETECTED")
                print("Creating baseline...")
                print()
                print("Baseline created successfully.")
            else:
                print("Comparison completed.")
            
            print()
            print(f"Current reviews: {len(current_reviews)}")
            print(f"New reviews: {len(comparison['new_reviews'])}")
            print(f"Existing reviews: {len(comparison['existing_reviews'])}")
            print(f"Potentially missing: {len(comparison['missing_reviews'])}")
            print(f"Returned: {len(comparison['returned_reviews'])}")
            
            if comparison['missing_reviews']:
                print()
                print("IMPORTANT:")
                print(f"{len(comparison['missing_reviews'])} reviews were previously seen but are not")
                print("currently returned by Google.")
                print("They are listed as \"Potentially Missing Reviews\".")
            
            print()
            print(f"Excel report:")
            print(f"{os.path.abspath(self.config['paths']['main_excel_file'])}")
            print()
            print("=" * 40)
            print("Completed successfully")
            print("=" * 40)
            
            # Log completion
            self.logger.info(f"Run completed successfully. Current reviews: {len(current_reviews)}, ")
            self.logger.info(f"New: {len(comparison['new_reviews'])}, Missing: {len(comparison['missing_reviews'])}")
            
        except Exception as e:
            print()
            print("=" * 40)
            print("ERROR OCCURRED")
            print("=" * 40)
            print()
            print(f"Error: {e}")
            print()
            print("Google review retrieval failed. Historical data was not modified.")
            print()
            print("=" * 40)
            
            self.logger.error(f"Run failed: {e}")
            raise


def main():
    """Main entry point."""
    import sys
    
    # Check for diagnostic mode
    if len(sys.argv) > 1 and sys.argv[1] == '--diagnose':
        try:
            backup = GoogleReviewsBackup()
            creds = backup._authenticate()
            backup._diagnose_api_access(creds)
            input("\nPress Enter to exit...")
        except Exception as e:
            print(f"\nDiagnostic error: {e}")
            input("\nPress Enter to exit...")
            exit(1)
        return
    
    try:
        backup = GoogleReviewsBackup()
        backup.run()
    except Exception as e:
        print(f"\nFatal error: {e}")
        input("\nPress Enter to exit...")
        exit(1)
    
    input("\nPress Enter to exit...")


if __name__ == '__main__':
    main()
