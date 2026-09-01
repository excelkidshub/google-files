import hashlib
import re
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generate_review_key(name, rating, review_message):
    """Generate SHA-256 hash from normalized review data."""
    normalized = f"{name.strip().lower()}|{rating}|{review_message.strip().lower()}"
    return hashlib.sha256(normalized.encode()).hexdigest()


def parse_rating(stars):
    """Convert star string to numeric rating."""
    star_count = stars.count('star')
    return star_count if 1 <= star_count <= 5 else 0


def parse_reviews_file(file_path):
    """Parse all-reviews.txt and extract review data."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    reviews = []
    lines = content.split('\n')
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Look for reviewer name line (contains "open_in_new")
        if 'open_in_new' in line:
            name = line.replace('open_in_new', '').strip()
            
            # Next line should contain review count and photos
            i += 1
            if i < len(lines):
                count_line = lines[i].strip()
            
            # Next line contains stars, date, and possibly status (all on one line)
            i += 1
            if i < len(lines):
                stars_date_line = lines[i].strip()
                rating = parse_rating(stars_date_line)
                
                # Extract date (e.g., "5 days ago", "1 week ago")
                date_match = re.search(r'\d+\s+(day|week|month)s?\s+ago', stars_date_line, re.IGNORECASE)
                review_date = date_match.group(0) if date_match else stars_date_line
                
                # Check for reply status
                reply_status = ''
                if 'New' in stars_date_line:
                    reply_status = 'New'
                elif 'Replied' in stars_date_line:
                    reply_status = 'Replied'
                elif 'Unreplied' in stars_date_line:
                    reply_status = 'Unreplied'
            
            # Next lines contain review message (may span multiple lines)
            i += 1
            review_message = ''
            while i < len(lines):
                next_line = lines[i].strip()
                # Stop if we hit another reviewer or empty line followed by reviewer pattern
                if 'open_in_new' in next_line:
                    break
                if not next_line and i + 1 < len(lines) and 'open_in_new' in lines[i + 1]:
                    i += 1
                    break
                if next_line:
                    if review_message:
                        review_message += ' ' + next_line
                    else:
                        review_message = next_line
                i += 1
            
            # Remove "View full review" suffix if present
            review_message = re.sub(r'\s*View full review\s*$', '', review_message).strip()
            
            # Only add if we have valid data (allow empty review message)
            if name and rating:
                # Use empty string if no message
                if not review_message:
                    review_message = ''
                review_key = generate_review_key(name, rating, review_message)
                reviews.append({
                    'Reviewer Name': name,
                    'Rating': rating,
                    'Review Message': review_message,
                    'Review Date': review_date,
                    'Reply Status': reply_status,
                    'Review Key': review_key
                })
            
            continue
        
        i += 1
    
    return reviews


def load_existing_reviews(excel_path):
    """Load existing reviews from All Reviews sheet."""
    if not Path(excel_path).exists():
        return {}
    
    wb = openpyxl.load_workbook(excel_path)
    
    # Try to load from 'All Reviews' sheet, fall back to active sheet
    if 'All Reviews' in wb.sheetnames:
        ws = wb['All Reviews']
    else:
        ws = wb.active
    
    existing = {}
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            continue  # Skip header
        
        if len(row) >= 6:
            review_key = row[0]  # Review Key is first column in All Reviews
            if review_key:
                existing[review_key] = {
                    'Review Key': review_key,
                    'Reviewer Name': row[1],
                    'Rating': row[2],
                    'Review Date': row[3],
                    'Review Message': row[4],
                    'Reply Status': row[5],
                    'First Seen': row[6] if len(row) > 6 else None,
                    'Last Seen': row[7] if len(row) > 7 else None,
                    'Status': row[8] if len(row) > 8 else None
                }
    
    wb.close()
    return existing


def setup_sheet(ws, headers):
    """Setup sheet with headers and formatting."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def save_reviews_to_excel(all_reviews, current_reviews, new_reviews, lost_reviews, duplicates, excel_path, today):
    """Save reviews to Excel with multiple sheets."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # === All Reviews Sheet ===
    ws_all = wb.create_sheet('All Reviews')
    headers_all = ['Review Key', 'Reviewer Name', 'Rating', 'Review Date', 'Review Message', 
                   'Reply Status', 'First Seen', 'Last Seen', 'Status']
    setup_sheet(ws_all, headers_all)
    
    for row_idx, review in enumerate(all_reviews, 2):
        ws_all.cell(row=row_idx, column=1, value=review['Review Key'])
        ws_all.cell(row=row_idx, column=2, value=review['Reviewer Name'])
        ws_all.cell(row=row_idx, column=3, value=review['Rating'])
        ws_all.cell(row=row_idx, column=4, value=review['Review Date'])
        ws_all.cell(row=row_idx, column=5, value=review['Review Message'])
        ws_all.cell(row=row_idx, column=6, value=review['Reply Status'])
        ws_all.cell(row=row_idx, column=7, value=review['First Seen'])
        ws_all.cell(row=row_idx, column=8, value=review['Last Seen'])
        ws_all.cell(row=row_idx, column=9, value=review['Status'])
    
    # Column widths for All Reviews
    ws_all.column_dimensions['A'].width = 40  # Review Key
    ws_all.column_dimensions['B'].width = 25  # Reviewer Name
    ws_all.column_dimensions['C'].width = 8   # Rating
    ws_all.column_dimensions['D'].width = 15  # Review Date
    ws_all.column_dimensions['E'].width = 60  # Review Message
    ws_all.column_dimensions['F'].width = 12  # Reply Status
    ws_all.column_dimensions['G'].width = 12  # First Seen
    ws_all.column_dimensions['H'].width = 12  # Last Seen
    ws_all.column_dimensions['I'].width = 15  # Status
    
    # Wrap text for Review Message
    for row in range(2, len(all_reviews) + 2):
        ws_all.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    
    # Color coding for status
    for row in range(2, len(all_reviews) + 2):
        status = ws_all.cell(row=row, column=9).value
        if status == 'NEW':
            ws_all.cell(row=row, column=9).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif status == 'LOST':
            ws_all.cell(row=row, column=9).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # === Current Reviews Sheet ===
    if current_reviews:
        ws_current = wb.create_sheet('Current')
        headers_current = ['Review Key', 'Reviewer Name', 'Rating', 'Review Date', 'Review Message', 'Reply Status']
        setup_sheet(ws_current, headers_current)
        
        for row_idx, review in enumerate(current_reviews, 2):
            ws_current.cell(row=row_idx, column=1, value=review['Review Key'])
            ws_current.cell(row=row_idx, column=2, value=review['Reviewer Name'])
            ws_current.cell(row=row_idx, column=3, value=review['Rating'])
            ws_current.cell(row=row_idx, column=4, value=review['Review Date'])
            ws_current.cell(row=row_idx, column=5, value=review['Review Message'])
            ws_current.cell(row=row_idx, column=6, value=review['Reply Status'])
        
        ws_current.column_dimensions['A'].width = 40
        ws_current.column_dimensions['B'].width = 25
        ws_current.column_dimensions['C'].width = 8
        ws_current.column_dimensions['D'].width = 15
        ws_current.column_dimensions['E'].width = 60
        ws_current.column_dimensions['F'].width = 12
        
        for row in range(2, len(current_reviews) + 2):
            ws_current.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    
    # === New Reviews Sheet ===
    if new_reviews:
        ws_new = wb.create_sheet('New')
        headers_new = ['Review Key', 'Reviewer Name', 'Rating', 'Review Date', 'Review Message', 'Reply Status', 'First Seen']
        setup_sheet(ws_new, headers_new)
        
        for row_idx, review in enumerate(new_reviews, 2):
            ws_new.cell(row=row_idx, column=1, value=review['Review Key'])
            ws_new.cell(row=row_idx, column=2, value=review['Reviewer Name'])
            ws_new.cell(row=row_idx, column=3, value=review['Rating'])
            ws_new.cell(row=row_idx, column=4, value=review['Review Date'])
            ws_new.cell(row=row_idx, column=5, value=review['Review Message'])
            ws_new.cell(row=row_idx, column=6, value=review['Reply Status'])
            ws_new.cell(row=row_idx, column=7, value=review['First Seen'])
        
        ws_new.column_dimensions['A'].width = 40
        ws_new.column_dimensions['B'].width = 25
        ws_new.column_dimensions['C'].width = 8
        ws_new.column_dimensions['D'].width = 15
        ws_new.column_dimensions['E'].width = 60
        ws_new.column_dimensions['F'].width = 12
        ws_new.column_dimensions['G'].width = 12
        
        for row in range(2, len(new_reviews) + 2):
            ws_new.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    
    # === Lost Reviews Sheet ===
    if lost_reviews:
        ws_lost = wb.create_sheet('Lost')
        headers_lost = ['Review Key', 'Reviewer Name', 'Rating', 'Review Date', 'Review Message', 'Reply Status', 'First Seen', 'Last Seen']
        setup_sheet(ws_lost, headers_lost)
        
        for row_idx, review in enumerate(lost_reviews, 2):
            ws_lost.cell(row=row_idx, column=1, value=review['Review Key'])
            ws_lost.cell(row=row_idx, column=2, value=review['Reviewer Name'])
            ws_lost.cell(row=row_idx, column=3, value=review['Rating'])
            ws_lost.cell(row=row_idx, column=4, value=review['Review Date'])
            ws_lost.cell(row=row_idx, column=5, value=review['Review Message'])
            ws_lost.cell(row=row_idx, column=6, value=review['Reply Status'])
            ws_lost.cell(row=row_idx, column=7, value=review['First Seen'])
            ws_lost.cell(row=row_idx, column=8, value=review['Last Seen'])
        
        ws_lost.column_dimensions['A'].width = 40
        ws_lost.column_dimensions['B'].width = 25
        ws_lost.column_dimensions['C'].width = 8
        ws_lost.column_dimensions['D'].width = 15
        ws_lost.column_dimensions['E'].width = 60
        ws_lost.column_dimensions['F'].width = 12
        ws_lost.column_dimensions['G'].width = 12
        ws_lost.column_dimensions['H'].width = 12
        
        for row in range(2, len(lost_reviews) + 2):
            ws_lost.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    
    # === Duplicates Sheet ===
    if duplicates:
        ws_dup = wb.create_sheet('Duplicates')
        headers_dup = ['Review Key', 'Reviewer Name', 'Rating', 'Occurrence Count']
        setup_sheet(ws_dup, headers_dup)
        
        for row_idx, dup in enumerate(duplicates, 2):
            ws_dup.cell(row=row_idx, column=1, value=dup['Review Key'])
            ws_dup.cell(row=row_idx, column=2, value=dup['Reviewer Name'])
            ws_dup.cell(row=row_idx, column=3, value=dup['Rating'])
            ws_dup.cell(row=row_idx, column=4, value=dup['Count'])
        
        ws_dup.column_dimensions['A'].width = 40
        ws_dup.column_dimensions['B'].width = 25
        ws_dup.column_dimensions['C'].width = 8
        ws_dup.column_dimensions['D'].width = 15
    
    wb.save(excel_path)
    wb.close()


def main():
    """Main execution function."""
    script_dir = Path(__file__).parent
    
    # Allow command-line argument for input file
    if len(sys.argv) > 1:
        input_file = Path(sys.argv[1])
    else:
        input_file = script_dir / 'all-reviews.txt'
    
    excel_file = script_dir / 'google_reviews.xlsx'
    
    # If file exists, use timestamped filename to avoid conflicts
    if excel_file.exists():
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = script_dir / f'google_reviews_{timestamp}.xlsx'
        print(f"Note: Existing file found, saving as: {excel_file.name}")
    today = datetime.now().strftime('%Y-%m-%d')
    
    print("Google Review Tracker")
    print("---------------------")
    print()
    
    # Parse input file
    if not input_file.exists():
        print(f"Error: {input_file} not found.")
        return
    
    parsed_reviews = parse_reviews_file(input_file)
    print(f"Parsed {len(parsed_reviews)} reviews from {input_file.name}")
    
    # Track duplicates in input
    key_counts = {}
    for review in parsed_reviews:
        key = review['Review Key']
        key_counts[key] = key_counts.get(key, 0) + 1
    
    # Get unique reviews from input (remove duplicates)
    unique_input_reviews = {}
    for review in parsed_reviews:
        key = review['Review Key']
        if key not in unique_input_reviews:
            unique_input_reviews[key] = review
    
    # Build duplicates list
    duplicates = []
    for key, count in key_counts.items():
        if count > 1:
            review = unique_input_reviews[key]
            duplicates.append({
                'Review Key': key,
                'Reviewer Name': review['Reviewer Name'],
                'Rating': review['Rating'],
                'Count': count
            })
    
    # Load existing reviews from All Reviews sheet
    existing_reviews = load_existing_reviews(excel_file)
    
    # COMPARISON LOGIC:
    # - Input key exists in historical data → CURRENT
    # - Input key does NOT exist in historical data → NEW
    # - Historical key does NOT exist in input → LOST
    
    input_keys = set(unique_input_reviews.keys())
    historical_keys = set(existing_reviews.keys())
    
    new_keys = input_keys - historical_keys  # In input, not in history → NEW
    current_keys = input_keys & historical_keys  # In both → CURRENT
    lost_keys = historical_keys - input_keys  # In history, not in input → LOST
    
    # Build All Reviews list
    all_reviews = []
    
    # Add NEW reviews
    for key in new_keys:
        review = unique_input_reviews[key]
        all_reviews.append({
            'Review Key': key,
            'Reviewer Name': review['Reviewer Name'],
            'Rating': review['Rating'],
            'Review Date': review['Review Date'],
            'Review Message': review['Review Message'],
            'Reply Status': review['Reply Status'],
            'First Seen': today,
            'Last Seen': today,
            'Status': 'NEW'
        })
    
    # Add CURRENT reviews (update Last Seen, keep First Seen)
    for key in current_keys:
        review = unique_input_reviews[key]
        historical = existing_reviews[key]
        all_reviews.append({
            'Review Key': key,
            'Reviewer Name': review['Reviewer Name'],
            'Rating': review['Rating'],
            'Review Date': review['Review Date'],
            'Review Message': review['Review Message'],
            'Reply Status': review['Reply Status'],
            'First Seen': historical['First Seen'],
            'Last Seen': today,
            'Status': 'CURRENT'
        })
    
    # Add LOST reviews (from historical, not in input)
    for key in lost_keys:
        historical = existing_reviews[key]
        all_reviews.append({
            'Review Key': key,
            'Reviewer Name': historical['Reviewer Name'],
            'Rating': historical['Rating'],
            'Review Date': historical['Review Date'],
            'Review Message': historical['Review Message'],
            'Reply Status': historical['Reply Status'],
            'First Seen': historical['First Seen'],
            'Last Seen': historical['Last Seen'],
            'Status': 'LOST'
        })
    
    # Build separate lists for sheets
    current_reviews_list = [unique_input_reviews[key] for key in current_keys]
    new_reviews_list = [unique_input_reviews[key] for key in new_keys]
    lost_reviews_list = [existing_reviews[key] for key in lost_keys]
    
    # Add First Seen to new reviews list for the New sheet
    for review in new_reviews_list:
        review['First Seen'] = today
    
    # Save to Excel with multiple sheets
    save_reviews_to_excel(all_reviews, current_reviews_list, new_reviews_list, lost_reviews_list, duplicates, excel_file, today)
    
    # Print summary
    print()
    print(f"Previous reviews (historical) : {len(existing_reviews)}")
    print(f"Current reviews (in input)   : {len(unique_input_reviews)}")
    print()
    print(f"NEW reviews      : {len(new_keys)}")
    print(f"CURRENT reviews  : {len(current_keys)}")
    print(f"LOST reviews     : {len(lost_keys)}")
    print(f"DUPLICATES       : {len(duplicates)}")
    print()
    
    if len(new_keys) == 0 and len(lost_keys) == 0 and len(duplicates) == 0:
        print("No new, lost, or duplicate reviews detected.")
    else:
        print("Excel updated:")
        print(f"  {excel_file.name}")
        print()
        print("Sheets created:")
        print("  - All Reviews (historical master list)")
        if current_reviews_list:
            print("  - Current (reviews in latest input)")
        if new_reviews_list:
            print("  - New (newly discovered reviews)")
        if lost_reviews_list:
            print("  - Lost (reviews missing from latest input)")
        if duplicates:
            print("  - Duplicates (duplicate occurrences in input)")


if __name__ == '__main__':
    main()
