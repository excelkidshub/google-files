from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "ExcelKidsHub-Master-Data.xlsx"
MAX_DATA_ROWS = 500

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, color="1F2937")
THIN_SIDE = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
DATE_FORMAT = "yyyy-mm-dd"
CURRENCY_FORMAT = u'\u20b9#,##0.00'
BATCH_SECTION_FILL = PatternFill("solid", fgColor="EAF4EA")
CARD_FILL = PatternFill("solid", fgColor="F8FAFC")
CARD_TITLE_FILL = PatternFill("solid", fgColor="DCEAF7")
CARD_TITLE_FONT = Font(bold=True, size=14, color="1F2937")
LABEL_FONT = Font(bold=True, color="334155")


def style_header_row(worksheet) -> None:
    """Apply professional styling to the header row."""
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = CENTER_ALIGNMENT


def apply_borders(worksheet) -> None:
    """Add borders to the used grid including pre-created formula rows."""
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = CELL_BORDER


def freeze_header(worksheet) -> None:
    """Freeze the top row for easier daily use."""
    worksheet.freeze_panes = "A2"


def auto_fit_columns(worksheet) -> None:
    """Auto-size columns based on the longest displayed value."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 28)


def set_column_formats(worksheet, date_columns: Iterable[str], currency_columns: Iterable[str]) -> None:
    """Apply date and currency formats to the target columns."""
    for column in date_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = DATE_FORMAT

    for column in currency_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = CURRENCY_FORMAT


def center_columns(worksheet, columns: Iterable[str]) -> None:
    """Center align selected columns for readability."""
    for column in columns:
        for cell in worksheet[column]:
            cell.alignment = CENTER_ALIGNMENT


def left_align_body(worksheet) -> None:
    """Default non-header content to left alignment unless overridden later."""
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = LEFT_ALIGNMENT


def add_dropdown(worksheet, cell_range: str, options: list[str]) -> None:
    """Attach a list dropdown to the requested range."""
    quoted = ",".join(options)
    validation = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    validation.prompt = "Please select a value from the list."
    validation.error = "Please select a valid option from the dropdown."
    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def setup_sheet_basics(worksheet) -> None:
    """Common styling shared by all worksheets."""
    style_header_row(worksheet)
    freeze_header(worksheet)
    left_align_body(worksheet)


def create_admissions_sheet(workbook: Workbook) -> None:
    """Create the main admissions sheet with formulas and dropdowns."""
    worksheet = workbook.active
    worksheet.title = "admissions"

    headers = [
        "Admission ID",
        "Parent Name",
        "Mobile",
        "Email",
        "Address",
        "City",
        "Student Name",
        "Age",
        "Gender",
        "School",
        "Grade",
        "Level",
        "Batch Code",
        "Mode",
        "Start Date",
        "End Date",
        "Status",
        "Total Fee",
        "Discount",
        "Manual Adjustment",
        "Adjusted Fee",
        "Total Paid",
        "Pending",
        "Payment Status",
        "Admission Source",
        "Referral Type",
        "Referrer Name",
        "Created Date",
        "Certificate Status",
        "Certificate Number",
        "Certificate Issue Date",
        "Certificate Sent Date",
    ]
    worksheet.append(headers)

    for row in range(2, MAX_DATA_ROWS + 2):
        worksheet[f"U{row}"] = f"=IF(OR(R{row}=\"\",S{row}=\"\",T{row}=\"\"),\"\",R{row}-S{row}+T{row})"
        worksheet[f"V{row}"] = f'=IF(A{row}="","",SUMIF(payments!B:B,A{row},payments!D:D))'
        worksheet[f"W{row}"] = f'=IF(U{row}="","",U{row}-V{row})'
        worksheet[f"X{row}"] = (
            f'=IF(U{row}="","",IF(V{row}=0,"Not Started",IF(V{row}<U{row},"Partial","Completed")))'
        )

    setup_sheet_basics(worksheet)
    add_dropdown(worksheet, f"L2:L{MAX_DATA_ROWS + 1}", ["Basic", "Advanced", "Proficient"])
    add_dropdown(worksheet, f"N2:N{MAX_DATA_ROWS + 1}", ["Online", "Offline"])
    add_dropdown(
        worksheet,
        f"Q2:Q{MAX_DATA_ROWS + 1}",
        ["Pending Start", "Active", "Completed", "Dropped"],
    )
    add_dropdown(worksheet, f"X2:X{MAX_DATA_ROWS + 1}", ["Not Started", "Partial", "Completed"])
    add_dropdown(worksheet, f"AC2:AC{MAX_DATA_ROWS + 1}", ["Not Issued", "Ready", "Sent"])

    set_column_formats(worksheet, ["O", "P", "AB", "AE", "AF"], ["R", "S", "T", "U", "V", "W"])
    center_columns(worksheet, ["A", "H", "I", "K", "L", "M", "N", "Q", "X"])
    apply_borders(worksheet)
    auto_fit_columns(worksheet)
    worksheet.column_dimensions["E"].width = 24
    worksheet.column_dimensions["F"].width = 16
    worksheet.column_dimensions["AC"].width = 16
    worksheet.column_dimensions["AD"].width = 22


def create_payments_sheet(workbook: Workbook) -> None:
    """Create the payments tracking sheet."""
    worksheet = workbook.create_sheet("payments")
    headers = [
        "Payment ID",
        "Admission ID",
        "Payment Date",
        "Amount Paid",
        "Payment Mode",
        "Transaction ID",
        "Notes",
    ]
    worksheet.append(headers)

    setup_sheet_basics(worksheet)
    set_column_formats(worksheet, ["C"], ["D"])
    center_columns(worksheet, ["A", "B", "E"])
    apply_borders(worksheet)
    auto_fit_columns(worksheet)


def create_batches_sheet(workbook: Workbook) -> None:
    """Create the batches sheet."""
    worksheet = workbook.create_sheet("batches")
    headers = ["Batch Code", "Level", "Mode", "Start Date", "End Date", "Timing", "Capacity"]
    worksheet.append(headers)

    setup_sheet_basics(worksheet)
    set_column_formats(worksheet, ["D", "E"], [])
    center_columns(worksheet, ["A", "B", "C", "G"])
    apply_borders(worksheet)
    auto_fit_columns(worksheet)


def create_expenses_sheet(workbook: Workbook) -> None:
    """Create the expenses sheet with category dropdown."""
    worksheet = workbook.create_sheet("expenses")
    headers = ["Date", "Category", "Sub Category", "Amount", "Payment Mode", "Notes"]
    worksheet.append(headers)

    setup_sheet_basics(worksheet)
    add_dropdown(
        worksheet,
        f"B2:B{MAX_DATA_ROWS + 1}",
        [
            "Salary",
            "Marketing",
            "Advertisement",
            "Printing",
            "Stationary",
            "Rent",
            "Utilities",
            "Software",
            "Miscellaneous",
        ],
    )
    set_column_formats(worksheet, ["A"], ["D"])
    center_columns(worksheet, ["B", "E"])
    apply_borders(worksheet)
    auto_fit_columns(worksheet)


def create_finance_summary_sheet(workbook: Workbook) -> None:
    """Create a richer monthly finance dashboard sheet."""
    worksheet = workbook.create_sheet("finance-summary")
    headers = [
        "Month",
        "Admission Count",
        "New Admissions Value",
        "Collections This Month",
        "Total Expense",
        "Profit",
        "Closing Pending",
        "Outstanding Admissions",
        "Salary Expense",
        "Marketing Expense",
        "Rent Expense",
        "Software Expense",
        "Other Expense",
    ]
    worksheet.append(headers)

    months = [f"2026-{month:02d}" for month in range(1, 13)]
    for row_index, month in enumerate(months, start=2):
        worksheet[f"A{row_index}"] = month
        worksheet[f"B{row_index}"] = (
            f'=COUNTIF(TEXT(admissions!$AB$2:$AB${MAX_DATA_ROWS + 1},"yyyy-mm"),$A{row_index})'
        )
        worksheet[f"C{row_index}"] = (
            f'=SUMPRODUCT((TEXT(admissions!$AB$2:$AB${MAX_DATA_ROWS + 1},"yyyy-mm")=$A{row_index})*(admissions!$U$2:$U${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"D{row_index}"] = (
            f'=SUMPRODUCT((TEXT(payments!$C$2:$C${MAX_DATA_ROWS + 1},"yyyy-mm")=$A{row_index})*(payments!$D$2:$D${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"E{row_index}"] = (
            f'=SUMPRODUCT((TEXT(expenses!$A$2:$A${MAX_DATA_ROWS + 1},"yyyy-mm")=$A{row_index})*(expenses!$D$2:$D${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"F{row_index}"] = f"=D{row_index}-E{row_index}"
        worksheet[f"G{row_index}"] = (
            f'=SUMPRODUCT((TEXT(admissions!$AB$2:$AB${MAX_DATA_ROWS + 1},"yyyy-mm")<=$A{row_index})*(admissions!$W$2:$W${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"H{row_index}"] = (
            f'=SUMPRODUCT((admissions!$W$2:$W${MAX_DATA_ROWS + 1}>0)*(TEXT(admissions!$AB$2:$AB${MAX_DATA_ROWS + 1},"yyyy-mm")<=$A{row_index}))'
        )
        worksheet[f"I{row_index}"] = (
            f'=SUMPRODUCT((TEXT(expenses!$A$2:$A${MAX_DATA_ROWS + 1},"yyyy-mm")=$A{row_index})*(expenses!$B$2:$B${MAX_DATA_ROWS + 1}="Salary")*(expenses!$D$2:$D${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"J{row_index}"] = (
            f'=SUMPRODUCT(((TEXT(expenses!$A$2:$A${MAX_DATA_ROWS + 1},"yyyy-mm")=$A{row_index})*((expenses!$B$2:$B${MAX_DATA_ROWS + 1}="Marketing")+(expenses!$B$2:$B${MAX_DATA_ROWS + 1}="Advertisement")))*(expenses!$D$2:$D${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"K{row_index}"] = (
            f'=SUMPRODUCT((TEXT(expenses!$A$2:$A${MAX_DATA_ROWS + 1},"yyyy-mm")=$A{row_index})*(expenses!$B$2:$B${MAX_DATA_ROWS + 1}="Rent")*(expenses!$D$2:$D${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"L{row_index}"] = (
            f'=SUMPRODUCT((TEXT(expenses!$A$2:$A${MAX_DATA_ROWS + 1},"yyyy-mm")=$A{row_index})*(expenses!$B$2:$B${MAX_DATA_ROWS + 1}="Software")*(expenses!$D$2:$D${MAX_DATA_ROWS + 1}))'
        )
        worksheet[f"M{row_index}"] = (
            f'=E{row_index}-I{row_index}-J{row_index}-K{row_index}-L{row_index}'
        )

    setup_sheet_basics(worksheet)
    set_column_formats(worksheet, [], ["C", "D", "E", "F", "G", "I", "J", "K", "L", "M"])
    center_columns(worksheet, ["A", "B", "H"])
    apply_borders(worksheet)
    auto_fit_columns(worksheet)


def style_card_range(worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """Apply card-like styling to a rectangular range."""
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.fill = CARD_FILL
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)


def create_batch_dashboard_sheet(workbook: Workbook) -> None:
    """Create a dashboard sheet with each batch shown as a visual card."""
    worksheet = workbook.create_sheet("batch-dashboard")
    worksheet.sheet_view.showGridLines = False

    card_height = 10
    row_gap = 2
    col_gap = 2
    card_width = 4
    left_starts = [1, 1 + card_width + col_gap]

    for row in range(1, 200):
        worksheet.row_dimensions[row].height = 22

    for column in ["A", "B", "C", "D", "G", "H", "I", "J"]:
        worksheet.column_dimensions[column].width = 16
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["J"].width = 20
    worksheet.column_dimensions["E"].width = 4
    worksheet.column_dimensions["F"].width = 4

    for batch_index in range(MAX_DATA_ROWS):
        batch_row = batch_index + 2
        grid_row = batch_index // 2
        grid_col = batch_index % 2
        start_row = 1 + grid_row * (card_height + row_gap)
        start_col = left_starts[grid_col]
        end_row = start_row + card_height - 1
        end_col = start_col + card_width - 1

        title_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}"
        worksheet.merge_cells(title_range)
        title_cell = worksheet.cell(start_row, start_col)
        title_cell.value = f'=IF(batches!A{batch_row}="","",batches!A{batch_row})'
        title_cell.font = CARD_TITLE_FONT
        title_cell.fill = CARD_TITLE_FILL
        title_cell.border = CELL_BORDER
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        style_card_range(worksheet, start_row + 1, end_row, start_col, end_col)
        for column_index in range(start_col, end_col + 1):
            worksheet.cell(start_row, column_index).fill = CARD_TITLE_FILL
            worksheet.cell(start_row, column_index).border = CELL_BORDER

        labels = [
            ("Level / Mode", f'=IF({get_column_letter(start_col)}{start_row}="","",batches!B{batch_row}&" / "&batches!C{batch_row})'),
            ("Dates", f'=IF({get_column_letter(start_col)}{start_row}="","",TEXT(batches!D{batch_row},"yyyy-mm-dd")&" to "&TEXT(batches!E{batch_row},"yyyy-mm-dd"))'),
            ("Duration", f'=IF(OR(batches!D{batch_row}="",batches!E{batch_row}=""),"",batches!E{batch_row}-batches!D{batch_row}+1&" days")'),
            ("Timing", f'=IF({get_column_letter(start_col)}{start_row}="","",batches!F{batch_row})'),
            ("Capacity", f'=IF({get_column_letter(start_col)}{start_row}="","",batches!G{batch_row})'),
            ("Enrolled", f'=IF({get_column_letter(start_col)}{start_row}="","",COUNTIF(admissions!$M:$M,batches!A{batch_row}))'),
            ("Available", f'=IF(OR(batches!A{batch_row}="",batches!G{batch_row}=""),"",batches!G{batch_row}-COUNTIF(admissions!$M:$M,batches!A{batch_row}))'),
        ]

        current_row = start_row + 1
        for label, formula in labels:
            worksheet.merge_cells(
                f"{get_column_letter(start_col + 1)}{current_row}:{get_column_letter(end_col)}{current_row}"
            )
            label_cell = worksheet.cell(current_row, start_col)
            label_cell.value = label
            label_cell.font = LABEL_FONT
            label_cell.fill = CARD_FILL
            label_cell.border = CELL_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell = worksheet.cell(current_row, start_col + 1)
            value_cell.value = formula
            value_cell.border = CELL_BORDER
            value_cell.fill = CARD_FILL
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        worksheet.merge_cells(
            f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{current_row}"
        )
        student_header = worksheet.cell(current_row, start_col)
        student_header.value = "Student List"
        student_header.font = LABEL_FONT
        student_header.fill = BATCH_SECTION_FILL
        student_header.border = CELL_BORDER
        student_header.alignment = Alignment(horizontal="left", vertical="center")

        worksheet.merge_cells(
            f"{get_column_letter(start_col)}{current_row + 1}:{get_column_letter(end_col)}{end_row}"
        )
        student_list_cell = worksheet.cell(current_row + 1, start_col)
        student_list_cell.value = (
            f'=IF(batches!A{batch_row}="","",TEXTJOIN(CHAR(10),TRUE,FILTER(admissions!$G$2:$G${MAX_DATA_ROWS + 1},admissions!$M$2:$M${MAX_DATA_ROWS + 1}=batches!A{batch_row},"")))'
        )
        student_list_cell.border = CELL_BORDER
        student_list_cell.fill = CARD_FILL
        student_list_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A1"



def create_workbook() -> Workbook:
    """Build the complete workbook with all required sheets."""
    workbook = Workbook()
    create_admissions_sheet(workbook)
    create_payments_sheet(workbook)
    create_batches_sheet(workbook)
    create_expenses_sheet(workbook)
    create_finance_summary_sheet(workbook)
    create_batch_dashboard_sheet(workbook)
    return workbook


def save_workbook(output_path: Path) -> None:
    """Create and save the workbook to disk."""
    workbook = create_workbook()
    workbook.save(output_path)


def main() -> None:
    """Script entry point."""
    output_path = Path(__file__).resolve().parent / OUTPUT_FILE
    save_workbook(output_path)
    print(f"Workbook created at: {output_path}")


if __name__ == "__main__":
    main()
